#!/usr/bin/env python3
import argparse
import json
import re
from pathlib import Path

import openpyxl


STAGE_SEQUENCE = [
    ("L1", "01"), ("L1", "02"), ("L1", "03"), ("L1", "04"), ("L1", "05"), ("L1", "06"), ("L1", "07"), ("L1", "08"), ("L1", "09"),
    ("L2", "01"), ("L2", "02"), ("L2", "03"), ("L2", "04"), ("L2", "05"), ("L2", "06"), ("L2", "07"), ("L2", "08"), ("L2", "09"),
    ("L3", "01"), ("L3", "02"), ("L3", "03"), ("L3", "04"), ("L3", "05"), ("L3", "06"), ("L3", "07"), ("L3", "08"), ("L3", "09"),
]

OPTION_LINE_RE = re.compile(r"^\s*([A-D])\s*[\.．、\)]\s*(.+?)\s*$", re.IGNORECASE)
ANSWER_RE = re.compile(r"[A-D]", re.IGNORECASE)


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\u00a0", " ").strip()


def stage_for_index(idx: int, total: int) -> tuple[str, str]:
    bucket = int((idx * len(STAGE_SEQUENCE)) / max(1, total))
    bucket = max(0, min(len(STAGE_SEQUENCE) - 1, bucket))
    return STAGE_SEQUENCE[bucket]


def parse_options(question_text: str) -> tuple[str, dict[str, str]]:
    lines = [line.strip() for line in question_text.replace("\r\n", "\n").split("\n") if line.strip()]
    stem_lines: list[str] = []
    options: dict[str, str] = {}
    for line in lines:
        m = OPTION_LINE_RE.match(line)
        if m:
            label = m.group(1).upper()
            text = m.group(2).strip()
            if text:
                options[label] = text
        else:
            stem_lines.append(line)
    stem = " ".join(stem_lines).strip()
    stem = re.sub(r"\s+", " ", stem)
    return stem, options


def normalize_answer(value: str) -> str:
    m = ANSWER_RE.search(clean_text(value))
    return m.group(0).upper() if m else ""


def build_distractor_pool(rows: list[dict]) -> list[str]:
    pool: list[str] = []
    seen = set()
    for row in rows:
        for text in row["options"].values():
            t = clean_text(text)
            if t and t not in seen:
                seen.add(t)
                pool.append(t)
    return pool


def fill_to_four_options(options: dict[str, str], correct: str, pool: list[str]) -> dict[str, str]:
    labels = ["A", "B", "C", "D"]
    result = {k: v for k, v in options.items() if k in labels and clean_text(v)}
    if correct not in result:
        return result
    existing_values = set(result.values())
    for label in labels:
        if label in result:
            continue
        candidate = ""
        for p in pool:
            if p not in existing_values and p != result[correct]:
                candidate = p
                break
        if candidate:
            result[label] = candidate
            existing_values.add(candidate)
    return {k: result[k] for k in labels if k in result}


def parse_workbook(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    entries: list[dict] = []

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [clean_text(h) for h in header_row]
        header_index = {name: i for i, name in enumerate(headers)}

        required = ["年级", "题目", "答案", "解析"]
        if any(name not in header_index for name in required):
            continue

        for row in ws.iter_rows(min_row=2, values_only=True):
            grade = clean_text(row[header_index["年级"]])
            question_raw = clean_text(row[header_index["题目"]])
            answer_raw = clean_text(row[header_index["答案"]])
            explanation = clean_text(row[header_index["解析"]])
            knowledge = clean_text(row[header_index.get("知识点", -1)]) if "知识点" in header_index else ""

            if not (grade and question_raw and answer_raw):
                continue

            stem, options = parse_options(question_raw)
            answer = normalize_answer(answer_raw)
            if not (stem and options and answer and answer in options):
                continue

            entries.append(
                {
                    "grade": grade,
                    "sheet": sheet,
                    "stem": stem,
                    "options": options,
                    "correct_answer": answer,
                    "explanation": explanation,
                    "knowledge": knowledge,
                }
            )

    return entries


def build_questions(entries: list[dict], prefix: str) -> list[dict]:
    pool = build_distractor_pool(entries)
    questions: list[dict] = []
    seen = set()
    total = len(entries)
    for idx, e in enumerate(entries):
        realm, stage = stage_for_index(idx, total)
        options = fill_to_four_options(e["options"], e["correct_answer"], pool)
        if e["correct_answer"] not in options:
            continue

        question_key = (e["stem"], json.dumps(options, ensure_ascii=False, sort_keys=True), e["correct_answer"])
        if question_key in seen:
            continue
        seen.add(question_key)

        question_id = f"{prefix}-{realm}-{stage}-{idx + 1:04d}"
        source_hint = f"来源：{e['sheet']} · {e['grade']}"
        if e["knowledge"]:
            source_hint += f" · 知识点：{e['knowledge']}"
        explanation = e["explanation"] or source_hint
        if e["explanation"]:
            explanation = f"{e['explanation']}（{source_hint}）"

        questions.append(
            {
                "question_id": question_id,
                "type": "grammar",
                "realm": realm,
                "stage": stage,
                "question": e["stem"],
                "options": options,
                "correct_answer": e["correct_answer"],
                "explanation": explanation,
                "word": "",
            }
        )
    return questions


def summarize(questions: list[dict]) -> dict:
    stage_counts: dict[str, int] = {}
    for q in questions:
        key = f"{q['realm']}-{q['stage']}"
        stage_counts[key] = stage_counts.get(key, 0) + 1
    return {
        "total_questions": len(questions),
        "stage_counts": dict(sorted(stage_counts.items())),
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Import elementary grammar xlsx files and export normalized JSON.")
    parser.add_argument("--xlsx", action="append", required=True, help="Path to xlsx file; repeat for multiple files")
    parser.add_argument("--output", default="database/data/elementary_grammar_questions.json", help="Output JSON path")
    parser.add_argument("--prefix", default="EG", help="Question ID prefix")
    args = parser.parse_args()

    xlsx_paths = [Path(p).expanduser() for p in args.xlsx]
    for p in xlsx_paths:
        if not p.exists():
            raise FileNotFoundError(f"Excel file not found: {p}")

    merged_entries: list[dict] = []
    for p in xlsx_paths:
        merged_entries.extend(parse_workbook(p))

    questions = build_questions(merged_entries, args.prefix)
    summary = summarize(questions)

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as f:
        json.dump(
            {
                "sources": [str(p) for p in xlsx_paths],
                "summary": summary,
                "questions": questions,
            },
            f,
            ensure_ascii=False,
            indent=2,
        )

    print(f"Exported {summary['total_questions']} questions -> {output_path}")
    print("Stage distribution:")
    for key, count in summary["stage_counts"].items():
        print(f"  {key}: {count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

