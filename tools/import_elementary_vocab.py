#!/usr/bin/env python3
import argparse
import json
import re
from pathlib import Path

import openpyxl


STAGE_SEQUENCE = [
    ("L1", "01"), ("L1", "02"), ("L1", "03"), ("L1", "04"), ("L1", "05"), ("L1", "06"), ("L1", "07"), ("L1", "08"), ("L1", "09"),
    ("L2", "01"), ("L2", "02"), ("L2", "03"), ("L2", "04"), ("L2", "05"), ("L2", "06"), ("L2", "07"), ("L2", "08"), ("L2", "09"),
    ("L3", "01"), ("L3", "02"), ("L3", "03"), ("L3", "04"), ("L3", "05"), ("L3", "06"), ("L3", "07"), ("L3", "08"), ("L3", "09"),
]

DISTRACTOR_STEPS = [37, 73, 109, 149, 191, 233, 277, 313, 359, 401]


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_word(word: str) -> str:
    return re.sub(r"\s+", " ", word).strip().lower()


def normalize_meaning(raw_meaning: str) -> str:
    text = clean_text(raw_meaning)
    text = re.sub(r"\s+", " ", text)
    text = text.replace("；", ";")
    parts = [part.strip() for part in re.split(r"[;,/|]", text) if part.strip()]
    return parts[0] if parts else text


def build_options(entries: list[dict], idx: int) -> tuple[dict, str]:
    total = len(entries)
    target = entries[idx]["meaning_primary"]
    distractors: list[str] = []

    for step in DISTRACTOR_STEPS:
        candidate_idx = (idx + step) % total
        candidate = entries[candidate_idx]["meaning_primary"]
        if candidate == target or candidate in distractors:
            continue
        distractors.append(candidate)
        if len(distractors) == 3:
            break

    if len(distractors) < 3:
        for j, item in enumerate(entries):
            candidate = item["meaning_primary"]
            if j == idx or candidate == target or candidate in distractors:
                continue
            distractors.append(candidate)
            if len(distractors) == 3:
                break

    if len(distractors) < 3:
        raise RuntimeError(f"Insufficient distractors for index={idx}, word={entries[idx]['word']}")

    correct_slot = idx % 4
    labels = ["A", "B", "C", "D"]
    choices = [None, None, None, None]
    choices[correct_slot] = target
    di = 0
    for ci in range(4):
        if choices[ci] is None:
            choices[ci] = distractors[di]
            di += 1

    options = {labels[i]: choices[i] for i in range(4)}
    return options, labels[correct_slot]


def stage_for_index(idx: int, total: int) -> tuple[str, str]:
    bucket = int((idx * len(STAGE_SEQUENCE)) / max(1, total))
    bucket = max(0, min(len(STAGE_SEQUENCE) - 1, bucket))
    return STAGE_SEQUENCE[bucket]


def parse_workbook(xlsx_path: Path, sheet_name: str | None) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [clean_text(h) for h in header_row]
    header_index = {name: i for i, name in enumerate(headers)}

    required = ["所属课本", "所属单元", "英文单词", "中文释义"]
    missing = [name for name in required if name not in header_index]
    if missing:
        raise RuntimeError(f"Excel 缺少必要列: {', '.join(missing)}")

    entries = []
    seen = set()
    valid_flag_col = header_index.get("校验标记")

    for row in ws.iter_rows(min_row=2, values_only=True):
        book = clean_text(row[header_index["所属课本"]])
        unit = clean_text(row[header_index["所属单元"]])
        word = clean_text(row[header_index["英文单词"]])
        meaning = clean_text(row[header_index["中文释义"]])
        valid_flag = clean_text(row[valid_flag_col]) if valid_flag_col is not None else ""

        if not (book and unit and word and meaning):
            continue
        if valid_flag and valid_flag != "正常":
            continue

        key = (book, unit, normalize_word(word))
        if key in seen:
            continue
        seen.add(key)

        entries.append(
            {
                "book": book,
                "unit": unit,
                "word": word,
                "meaning": meaning,
                "meaning_primary": normalize_meaning(meaning),
            }
        )

    entries.sort(key=lambda x: (x["book"], x["unit"], normalize_word(x["word"])))
    return entries


def build_questions(entries: list[dict]) -> list[dict]:
    total = len(entries)
    questions = []
    for idx, entry in enumerate(entries):
        realm, stage = stage_for_index(idx, total)
        options, correct_answer = build_options(entries, idx)
        question_id = f"EV-{realm}-{stage}-{idx + 1:04d}"
        question_text = f"\"{entry['word']}\" 的中文意思是？"
        explanation = f"{entry['word']} = {entry['meaning']}（来源：{entry['book']} {entry['unit']}）"

        questions.append(
            {
                "question_id": question_id,
                "type": "vocab",
                "realm": realm,
                "stage": stage,
                "question": question_text,
                "options": options,
                "correct_answer": correct_answer,
                "explanation": explanation,
                "word": entry["word"],
            }
        )
    return questions


def summarize(questions: list[dict]) -> dict:
    stage_counts = {}
    for q in questions:
        key = f"{q['realm']}-{q['stage']}"
        stage_counts[key] = stage_counts.get(key, 0) + 1
    return {
        "total_questions": len(questions),
        "stage_counts": dict(sorted(stage_counts.items())),
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Import elementary vocab Excel and export JSON for levelup_questions.")
    parser.add_argument("--xlsx", required=True, help="Path to xlsx file")
    parser.add_argument("--output", default="database/data/elementary_vocab_questions.json", help="Output JSON path")
    parser.add_argument("--sheet", default="", help="Optional sheet name")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx).expanduser()
    output_path = Path(args.output)

    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")

    entries = parse_workbook(xlsx_path, args.sheet or None)
    questions = build_questions(entries)
    summary = summarize(questions)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as f:
        json.dump(
            {
                "source": str(xlsx_path),
                "sheet": args.sheet or "",
                "summary": summary,
                "questions": questions,
            },
            f,
            ensure_ascii=False,
            indent=2,
        )

    print(f"Exported {summary['total_questions']} questions -> {output_path}")
    print("Stage distribution:")
    for key, count in summary["stage_counts"].items():
        print(f"  {key}: {count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
